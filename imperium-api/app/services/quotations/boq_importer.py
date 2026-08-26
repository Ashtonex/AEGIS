import csv
import math
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Any
from io import BytesIO, StringIO
from openpyxl import load_workbook
from app.services.quotations.calculator import BOQItem


class BOQImportResult:
    def __init__(
        self, items: List[BOQItem], warnings: List[str], summary: Dict[str, Any]
    ):
        self.items = items
        self.warnings = warnings
        self.summary = summary

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": len(self.items) > 0,
            "items": [item.model_dump() for item in self.items],
            "warnings": self.warnings,
            "summary": self.summary,
        }


class BOQImporter:
    @staticmethod
    def _is_blank(val: Any) -> bool:
        if val is None:
            return True
        if isinstance(val, float) and math.isnan(val):
            return True
        return str(val).strip() == ""

    @staticmethod
    def _sanitize_decimal(val: Any) -> tuple[Decimal, bool]:
        """Returns (value, parse_failed). parse_failed is only True when the
        cell had content that wasn't a valid number - never for a genuinely
        empty/NaN cell, so callers can warn without flagging every blank cell."""
        if BOQImporter._is_blank(val):
            return Decimal("0"), False
        try:
            # Strip formatting characters like currency symbols or commas
            clean_str = str(val).replace("$", "").replace(",", "").strip()
            # Accounting-style negatives, e.g. "(1,234.56)" -> -1234.56
            if clean_str.startswith("(") and clean_str.endswith(")"):
                clean_str = "-" + clean_str[1:-1]
            d = Decimal(clean_str)
            if d.is_nan():
                return Decimal("0"), True
            return d, False
        except Exception:
            return Decimal("0"), True

    @staticmethod
    def _clean_text(val: Any) -> str:
        if BOQImporter._is_blank(val):
            return ""
        return str(val).strip()

    @classmethod
    def _import_excel_workbook(cls, file_content: bytes) -> BOQImportResult:
        workbook = load_workbook(BytesIO(file_content), data_only=False, read_only=True)
        warnings: List[str] = []
        items: List[BOQItem] = []
        total_direct_costs = Decimal("0")
        rows_processed = 0
        sections_seen: set[str] = set()

        for ws in workbook.worksheets:
            current_section = ws.title.strip() or "Measured Works"
            header_map: Dict[str, int] = {}
            pending_description = ""
            sections_seen.add(current_section)

            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = list(row)
                rows_processed += 1
                normalized = [cls._clean_text(v).lower() for v in values]
                joined = " ".join(v for v in normalized if v)

                if not any(normalized):
                    pending_description = ""
                    continue

                if "description" in joined or "details" in joined:
                    for idx, value in enumerate(normalized):
                        if value in {"item", "item no", "item no.", "no"}:
                            header_map["item_no"] = idx
                        elif "description" in value or "details" in value or value == "item":
                            header_map["description"] = idx
                        elif value in {"unit", "uom"}:
                            header_map["unit"] = idx
                        elif "quant" in value or value in {"qty", "quantity"}:
                            header_map["quantity"] = idx
                        elif "rate" in value or "unit rate" in value:
                            header_map["rate"] = idx
                    pending_description = ""
                    continue

                desc_idx = header_map.get("description", 1 if len(values) > 1 else 0)
                item_idx = header_map.get("item_no", 0)
                unit_idx = header_map.get("unit", 2)
                qty_idx = header_map.get("quantity", 4 if len(values) > 4 else 3)
                rate_idx = header_map.get("rate", 5 if len(values) > 5 else 4)

                desc = cls._clean_text(values[desc_idx] if desc_idx < len(values) else "")
                item_no = cls._clean_text(values[item_idx] if item_idx < len(values) else "")
                unit = cls._clean_text(values[unit_idx] if unit_idx < len(values) else "")
                raw_qty = values[qty_idx] if qty_idx < len(values) else None
                raw_rate = values[rate_idx] if rate_idx < len(values) else None
                qty, qty_parse_failed = cls._sanitize_decimal(raw_qty)
                rate, rate_parse_failed = cls._sanitize_decimal(raw_rate)

                has_measure = unit or qty != 0 or rate != 0
                if desc and not has_measure:
                    heading = desc.upper()
                    heading_like = desc == heading
                    if heading_like:
                        current_section = heading
                        sections_seen.add(current_section)
                    elif pending_description or item_no:
                        pending_description = f"{pending_description} {desc}".strip()
                    elif items:
                        last_item = items[-1]
                        last_item.description = f"{last_item.description} {desc}".strip()
                    else:
                        pending_description = f"{pending_description} {desc}".strip()
                    continue

                if not desc and pending_description and has_measure:
                    desc = pending_description
                    pending_description = ""
                elif desc and pending_description and has_measure:
                    desc = f"{pending_description} {desc}".strip()
                    pending_description = ""

                if not desc or "total carried" in desc.lower() or "final summary" in desc.lower():
                    continue

                if qty_parse_failed:
                    warnings.append(f"{ws.title} row {row_number}: Could not parse quantity '{raw_qty}' as a number; treated as 0.")
                if rate_parse_failed:
                    warnings.append(f"{ws.title} row {row_number}: Could not parse rate '{raw_rate}' as a number; treated as 0.")
                if qty < 0:
                    warnings.append(f"{ws.title} row {row_number}: Negative quantity ({qty}) set to 0.")
                    qty = Decimal("0")
                if rate < 0:
                    warnings.append(f"{ws.title} row {row_number}: Negative rate ({rate}) set to 0.")
                    rate = Decimal("0")

                item = BOQItem(
                    section=current_section,
                    item_no=item_no,
                    description=desc,
                    quantity=qty,
                    unit=unit or "item",
                    rate=rate,
                )
                items.append(item)
                total_direct_costs += (qty * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        return BOQImportResult(
            items=items,
            warnings=warnings,
            summary={
                "total_rows_processed": rows_processed,
                "valid_items_imported": len(items),
                "section_count": len(sections_seen),
                "total_direct_costs": str(total_direct_costs),
            },
        )

    @classmethod
    def import_boq(cls, file_content: bytes, file_extension: str) -> BOQImportResult:
        """
        Parses BOQ items from an Excel or CSV file.
        Dynamically maps column headers to match 'description', 'quantity', 'unit', and 'rate'.
        """
        warnings = []
        items: List[BOQItem] = []

        try:
            if file_extension.lower() in [".xlsx", ".xls"]:
                return cls._import_excel_workbook(file_content)
            elif file_extension.lower() == ".csv":
                text = file_content.decode("utf-8-sig", errors="replace")
                rows = list(csv.DictReader(StringIO(text)))
                columns = [str(c).strip().lower() for c in (rows[0].keys() if rows else [])]
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
        except Exception as e:
            return BOQImportResult(
                items=[],
                warnings=[f"Failed to read file: {str(e)}"],
                summary={
                    "total_rows_processed": 0,
                    "valid_items_imported": 0,
                    "total_direct_costs": "0.00",
                },
            )

        # Column mapping matrix
        desc_cols = [
            "description",
            "item description",
            "task",
            "details",
            "name",
            "item",
        ]
        qty_cols = ["quantity", "qty", "volume", "amount_qty"]
        unit_cols = ["unit", "uom", "measure"]
        rate_cols = ["rate", "unit rate", "price", "unit price", "cost"]
        section_cols = ["section", "bill", "trade", "heading", "category"]
        item_no_cols = ["item no", "item no.", "item_no", "item", "no"]

        # Helper to find first matching column
        def find_col(possible_names: List[str], fallback: str) -> str:
            for col in columns:
                if col in possible_names:
                    return col
            return fallback

        desc_col = find_col(desc_cols, "description")
        qty_col = find_col(qty_cols, "quantity")
        unit_col = find_col(unit_cols, "unit")
        rate_col = find_col(rate_cols, "rate")
        section_col = find_col(section_cols, "section")
        item_no_col = find_col(item_no_cols, "item_no")

        if desc_col not in columns:
            warnings.append(
                "Could not find description column. Using first column as fallback."
            )
            desc_col = columns[0] if columns else "description"

        total_direct_costs = Decimal("0")
        rows_processed = 0

        for idx, row in enumerate(rows):
            rows_processed += 1
            # Retrieve values safely
            normalized_row = {str(k).strip().lower(): v for k, v in row.items()}
            raw_desc = normalized_row.get(desc_col) if desc_col in columns else None
            raw_qty = normalized_row.get(qty_col) if qty_col in columns else None
            raw_unit = normalized_row.get(unit_col) if unit_col in columns else None
            raw_rate = normalized_row.get(rate_col) if rate_col in columns else None
            raw_section = normalized_row.get(section_col) if section_col in columns else None
            raw_item_no = normalized_row.get(item_no_col) if item_no_col in columns else None

            # Skip completely empty rows
            if cls._is_blank(raw_desc) and cls._is_blank(raw_qty) and cls._is_blank(raw_rate):
                continue

            desc = cls._clean_text(raw_desc)
            if not desc:
                warnings.append(f"Row {idx + 1}: Empty description, skipping row.")
                continue

            qty, qty_parse_failed = cls._sanitize_decimal(raw_qty)
            unit = cls._clean_text(raw_unit) or "item"
            rate, rate_parse_failed = cls._sanitize_decimal(raw_rate)

            if qty_parse_failed:
                warnings.append(f"Row {idx + 1}: Could not parse quantity '{raw_qty}' as a number; treated as 0.")
            if rate_parse_failed:
                warnings.append(f"Row {idx + 1}: Could not parse rate '{raw_rate}' as a number; treated as 0.")

            if qty < 0:
                warnings.append(f"Row {idx + 1}: Negative quantity ({qty}) set to 0.")
                qty = Decimal("0")
            if rate < 0:
                warnings.append(f"Row {idx + 1}: Negative rate ({rate}) set to 0.")
                rate = Decimal("0")

            section = (cls._clean_text(raw_section) if section_col in columns else "") or "Measured Works"
            item_no = cls._clean_text(raw_item_no) if item_no_col in columns else ""
            boq_item = BOQItem(section=section, item_no=item_no, description=desc, quantity=qty, unit=unit, rate=rate)
            items.append(boq_item)

            item_cost = (qty * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            total_direct_costs += item_cost

        summary = {
            "total_rows_processed": rows_processed,
            "valid_items_imported": len(items),
            "total_direct_costs": str(total_direct_costs),
        }

        return BOQImportResult(items=items, warnings=warnings, summary=summary)
