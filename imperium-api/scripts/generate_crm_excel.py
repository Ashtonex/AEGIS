from __future__ import annotations

import os
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DESKTOP_EXCEL_PATH = Path(r"C:\Users\ashjx\Desktop\CCB_Construction_CRM_Master.xlsx")
ARTIFACT_DIR = Path(r"C:\Users\ashjx\.gemini\antigravity-cli\brain\97cf26ac-1312-46eb-be2b-5d2fa1954a8f")
ARTIFACT_EXCEL_PATH = ARTIFACT_DIR / "CCB_Construction_CRM_Master.xlsx"


def create_crm_workbook():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    gold_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
    accent_blue_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    gray_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

    green_alert_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    amber_alert_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_alert_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="E2E8F0")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    card_title_font = Font(name="Calibri", size=9, bold=True, color="475569")
    card_value_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    regular_font = Font(name="Calibri", size=10, color="0F172A")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    thick_border_side = Side(border_style="medium", color="0F172A")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    bottom_double_border = Border(top=thin_border_side, bottom=Side(border_style="double", color="0F172A"))

    # ---------------------------------------------------------
    # TAB 1: EXECUTIVE DASHBOARD & CRM KPI OVERVIEW
    # ---------------------------------------------------------
    ws1 = wb.create_sheet(title="CRM Executive Dashboard")
    ws1.views.sheetView[0].showGridLines = True

    # Title Block
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "SIX NINE CONSTRUCTION (PVT) LTD — COMMERCIAL CONTROL BRAIN (CCB) CRM DASHBOARD"
    ws1["A1"].font = title_font
    ws1["A1"].fill = navy_fill
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws1.merge_cells("A2:J2")
    ws1["A2"] = "Executive Pipeline Management, Quotation Control & Commercial Guard Governance Engine"
    ws1["A2"].font = subtitle_font
    ws1["A2"].fill = navy_fill
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 28
    ws1.row_dimensions[2].height = 18

    # KPI Summary Cards (Row 4-5)
    kpis = [
        ("TOTAL PIPELINE VALUE", "='Client & Lead Master'!G14", "A", "B", "$#,##0.00"),
        ("TOTAL CONTRACTED VALUE", "='Quotations & BOQ Estimator'!K10", "D", "E", "$#,##0.00"),
        ("PROTECTED GROSS PROFIT", "='Quotations & BOQ Estimator'!I10", "G", "H", "$#,##0.00"),
        ("AVG PROTECTED MARGIN", "='Quotations & BOQ Estimator'!J10", "J", "J", "0.0%"),
    ]

    ws1.row_dimensions[4].height = 16
    ws1.row_dimensions[5].height = 26

    # KPI Card 1
    ws1.merge_cells("A4:B4")
    ws1["A4"] = "TOTAL ACTIVE PIPELINE"
    ws1["A4"].font = card_title_font
    ws1["A4"].fill = light_blue_fill
    ws1["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells("A5:B5")
    ws1["A5"] = "=SUM('Client & Lead Master'!G5:G14)"
    ws1["A5"].font = card_value_font
    ws1["A5"].fill = light_blue_fill
    ws1["A5"].number_format = "$#,##0.00"
    ws1["A5"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI Card 2
    ws1.merge_cells("D4:E4")
    ws1["D4"] = "TOTAL CONTRACTED QUOTES"
    ws1["D4"].font = card_title_font
    ws1["D4"].fill = light_blue_fill
    ws1["D4"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells("D5:E5")
    ws1["D5"] = "=SUM('Quotations & BOQ Estimator'!K5:K9)"
    ws1["D5"].font = card_value_font
    ws1["D5"].fill = light_blue_fill
    ws1["D5"].number_format = "$#,##0.00"
    ws1["D5"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI Card 3
    ws1.merge_cells("G4:H4")
    ws1["G4"] = "PROTECTED GROSS PROFIT"
    ws1["G4"].font = card_title_font
    ws1["G4"].fill = light_blue_fill
    ws1["G4"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells("G5:H5")
    ws1["G5"] = "=SUM('Quotations & BOQ Estimator'!I5:I9)"
    ws1["G5"].font = card_value_font
    ws1["G5"].fill = light_blue_fill
    ws1["G5"].number_format = "$#,##0.00"
    ws1["G5"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI Card 4
    ws1.merge_cells("J4:J4")
    ws1["J4"] = "CCB INTERCEPTED AUDITS"
    ws1["J4"].font = card_title_font
    ws1["J4"].fill = light_blue_fill
    ws1["J4"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["J5"] = "=COUNTA('CCB Guard Audits'!A5:A12)"
    ws1["J5"].font = card_value_font
    ws1["J5"].fill = light_blue_fill
    ws1["J5"].alignment = Alignment(horizontal="center", vertical="center")

    # Pipeline Funnel Table
    ws1["A7"] = "CRM Pipeline Funnel & Stage Breakdown"
    ws1["A7"].font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    ws1.row_dimensions[8].height = 22

    headers_funnel = ["Deal Stage", "Lead Count", "Total Value ($)", "Avg Win Prob (%)", "Weighted Pipeline Forecast ($)", "Target Gross Profit ($)"]
    for c_idx, h in enumerate(headers_funnel, 1):
        cell = ws1.cell(row=8, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = gray_header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")

    stages_data = [
        ("1. Lead / Initial Contact", 3, 185000, 0.10),
        ("2. Qualified & Site Visit", 2, 140000, 0.30),
        ("3. BOQ Takeoff & Costing", 2, 70000, 0.50),
        ("4. Proposal Sent / CCB Pack", 2, 95000, 0.75),
        ("5. Contract Signed / Active", 1, 45000, 1.00),
    ]

    for r_idx, (stage, count, val, prob) in enumerate(stages_data, 9):
        ws1.cell(row=r_idx, column=1, value=stage).font = regular_font
        ws1.cell(row=r_idx, column=2, value=count).font = regular_font
        ws1.cell(row=r_idx, column=2).alignment = Alignment(horizontal="center")

        c_val = ws1.cell(row=r_idx, column=3, value=val)
        c_val.font = regular_font
        c_val.number_format = "$#,##0.00"

        c_prob = ws1.cell(row=r_idx, column=4, value=prob)
        c_prob.font = regular_font
        c_prob.number_format = "0.0%"
        c_prob.alignment = Alignment(horizontal="center")

        c_weight = ws1.cell(row=r_idx, column=5, value=f"=C{r_idx}*D{r_idx}")
        c_weight.font = bold_font
        c_weight.number_format = "$#,##0.00"

        c_prof = ws1.cell(row=r_idx, column=6, value=f"=E{r_idx}*0.165")
        c_prof.font = regular_font
        c_prof.number_format = "$#,##0.00"

        for col in range(1, 7):
            ws1.cell(row=r_idx, column=col).border = thin_border

    # Totals Row
    tot_row = 14
    ws1.cell(row=tot_row, column=1, value="Total Active CRM Pipeline").font = bold_font
    ws1.cell(row=tot_row, column=2, value="=SUM(B9:B13)").font = bold_font
    ws1.cell(row=tot_row, column=2).alignment = Alignment(horizontal="center")
    ws1.cell(row=tot_row, column=3, value="=SUM(C9:C13)").font = bold_font
    ws1.cell(row=tot_row, column=3).number_format = "$#,##0.00"
    ws1.cell(row=tot_row, column=4, value="=").font = bold_font
    ws1.cell(row=tot_row, column=5, value="=SUM(E9:E13)").font = bold_font
    ws1.cell(row=tot_row, column=5).number_format = "$#,##0.00"
    ws1.cell(row=tot_row, column=6, value="=SUM(F9:F13)").font = bold_font
    ws1.cell(row=tot_row, column=6).number_format = "$#,##0.00"
    for col in range(1, 7):
        ws1.cell(row=tot_row, column=col).border = bottom_double_border

    # ---------------------------------------------------------
    # TAB 2: CLIENT & LEAD MASTER REGISTRY
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="Client & Lead Master")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:N1")
    ws2["A1"] = "CLIENT & LEAD MASTER REGISTRY"
    ws2["A1"].font = title_font
    ws2["A1"].fill = navy_fill
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 26

    headers_leads = [
        "Client ID", "Client / Company Name", "Contact Person", "Phone", "Email",
        "Project Type", "Target Budget ($)", "Deal Stage", "Win Prob (%)",
        "Weighted Value ($)", "Lead Source", "Assigned AE", "Est. Close Date", "Action / Status"
    ]
    for c_idx, h in enumerate(headers_leads, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[4].height = 22

    clients_data = [
        ("CL-101", "ZimRe Real Estate Investment Trust", "Tariro Moyo", "+263 772 112233", "tmoyo@zimre.co.zw", "Duplex Extension & Alterations", 45000, "5. Contract Signed / Active", 1.00, "Inbound Tender", "Farai Chitepo", "2026-07-15", "Active Site Possession"),
        ("CL-102", "Old Mutual Property Zimbabwe", "Blessing Ndlovu", "+263 773 445566", "bndlovu@oldmutual.co.zw", "Commercial Office Refurbishment", 120000, "4. Proposal Sent / CCB Pack", 0.75, "Key Account", "Farai Chitepo", "2026-08-01", "CCB 55-Page Pack Delivered"),
        ("CL-103", "First Mutual Properties", "Kudzai Samuriwo", "+263 771 889900", "ksamuriwo@firstmutual.co.zw", "Residential Cluster Housing", 250000, "3. BOQ Takeoff & Costing", 0.50, "Direct Referral", "Tendai Mudekwa", "2026-08-15", "AI Vision Plan Takeoff in Progress"),
        ("CL-104", "NSSA Real Estate Portfolio", "Simbarashe Gumbo", "+263 774 223344", "sgumbo@nssa.org.zw", "Industrial Warehouse Mezzanine", 85000, "2. Qualified & Site Visit", 0.30, "Government RFP", "Farai Chitepo", "2026-08-30", "Site Assessment Completed"),
        ("CL-105", "Dawn Properties Ltd", "Chipo Mutasa", "+263 775 667788", "cmutasa@dawnprops.co.zw", "Boutique Lodge Renovation", 65000, "1. Lead / Initial Contact", 0.10, "Website Inquiry", "Tendai Mudekwa", "2026-09-15", "Initial Kickoff Call Scheduled"),
        ("CL-106", "Highland Park Retail Development", "Rangarirai Zhou", "+263 776 990011", "rzhou@highlandpark.co.zw", "Retail Store Fitout", 35000, "4. Proposal Sent / CCB Pack", 0.75, "Walk-in Lead", "Farai Chitepo", "2026-07-28", "Client Reviewing Quotation"),
        ("CL-107", "Borrowdale Brooke Estate Res-09", "Dr. Kelvin Sibanda", "+263 777 334455", "ksibanda@medical.co.zw", "Luxury Villa Extension", 95000, "1. Lead / Initial Contact", 0.10, "Private Client Referral", "Tendai Mudekwa", "2026-09-30", "Architectural Drawings Ingested"),
        ("CL-108", "Graniteside Logistics Park", "Nigel Thornton", "+263 778 556677", "nthornton@graniteside.co.zw", "Boundary Wall & Paving Package", 25000, "5. Contract Signed / Active", 1.00, "Repeat Client", "Farai Chitepo", "2026-07-10", "Site Operations Live"),
        ("CL-109", "Msasa Industrial Holdings", "Tinashe Mapfumo", "+263 779 114477", "tmapfumo@msasaind.co.zw", "Factory Floor Screed & Epoxy", 55000, "2. Qualified & Site Visit", 0.30, "Trade Exhibition", "Tendai Mudekwa", "2026-08-20", "Subcontractor Quotes Issued"),
        ("CL-110", "Avondale Medical Chambers", "Dr. Ayesha Patel", "+263 771 992255", "apatel@avondalemed.co.zw", "Clinic Extension & Structural Slab", 75000, "1. Lead / Initial Contact", 0.10, "Direct Outreach", "Farai Chitepo", "2026-10-05", "Budget Consultation Required"),
    ]

    for r_idx, row_data in enumerate(clients_data, 5):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            if c_idx == 7: # Budget
                cell.value = val
                cell.number_format = "$#,##0.00"
                cell.font = bold_font
            elif c_idx == 9: # Win Prob
                cell.value = val
                cell.number_format = "0.0%"
                cell.font = regular_font
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 10: # Weighted Formula
                cell.value = f"=G{r_idx}*I{r_idx}"
                cell.number_format = "$#,##0.00"
                cell.font = bold_font
            else:
                cell.value = val
                cell.font = regular_font
            cell.border = thin_border

    # ---------------------------------------------------------
    # TAB 3: QUOTATIONS & BOQ ESTIMATOR
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="Quotations & BOQ Estimator")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:M1")
    ws3["A1"] = "QUOTATIONS & BOQ ESTIMATION CONTROL ENGINE"
    ws3["A1"].font = title_font
    ws3["A1"].fill = navy_fill
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[1].height = 26

    headers_quotes = [
        "Quote Ref", "Client Name", "Project Title", "GFA (m²)",
        "Direct Works Cost ($)", "Contingency ($)", "Management Reserve ($)",
        "Internal Cost Ceiling ($)", "Protected Gross Profit ($)", "Margin %",
        "Client Contract Value ($)", "AI Vision Takeoff Status", "CCB Status"
    ]
    for c_idx, h in enumerate(headers_quotes, 1):
        cell = ws3.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[4].height = 22

    quotes_data = [
        ("SNC-RES-45K-001", "ZimRe Real Estate Investment Trust", "Residential Duplex Extension", 220, 33000.00, 2500.00, 2075.00, 45000.00, "98.9% Confident (Verified)", "APPROVED & LOCKED"),
        ("SNC-COM-120K-002", "Old Mutual Property Zimbabwe", "Commercial Office Refurbishment", 550, 88000.00, 6000.00, 5800.00, 120000.00, "97.5% Confident (Verified)", "PROPOSAL SENT"),
        ("SNC-RES-250K-003", "First Mutual Properties", "Residential Cluster Housing", 1200, 185000.00, 12500.00, 11250.00, 250000.00, "96.8% Confident (Verified)", "IN REVIEW"),
        ("SNC-IND-85K-004", "NSSA Real Estate Portfolio", "Industrial Warehouse Mezzanine", 410, 62000.00, 4500.00, 4475.00, 85000.00, "99.1% Confident (Verified)", "QUALIFIED"),
        ("SNC-RET-35K-005", "Highland Park Retail Development", "Retail Store Fitout", 180, 25500.00, 2000.00, 1725.00, 35000.00, "98.2% Confident (Verified)", "PROPOSAL SENT"),
    ]

    for r_idx, q in enumerate(quotes_data, 5):
        ref, client, title, gfa, direct, cont, res, total_contract, vision_stat, ccb_stat = q

        ws3.cell(row=r_idx, column=1, value=ref).font = bold_font
        ws3.cell(row=r_idx, column=2, value=client).font = regular_font
        ws3.cell(row=r_idx, column=3, value=title).font = regular_font

        c_gfa = ws3.cell(row=r_idx, column=4, value=gfa)
        c_gfa.font = regular_font
        c_gfa.alignment = Alignment(horizontal="center")

        c_dir = ws3.cell(row=r_idx, column=5, value=direct)
        c_dir.font = regular_font
        c_dir.number_format = "$#,##0.00"

        c_cont = ws3.cell(row=r_idx, column=6, value=cont)
        c_cont.font = regular_font
        c_cont.number_format = "$#,##0.00"

        c_res = ws3.cell(row=r_idx, column=7, value=res)
        c_res.font = regular_font
        c_res.number_format = "$#,##0.00"

        # Formulas
        c_ceil = ws3.cell(row=r_idx, column=8, value=f"=E{r_idx}+F{r_idx}+G{r_idx}")
        c_ceil.font = bold_font
        c_ceil.number_format = "$#,##0.00"

        c_prof = ws3.cell(row=r_idx, column=9, value=f"=K{r_idx}-H{r_idx}")
        c_prof.font = bold_font
        c_prof.number_format = "$#,##0.00"

        c_marg = ws3.cell(row=r_idx, column=10, value=f"=I{r_idx}/K{r_idx}")
        c_marg.font = bold_font
        c_marg.number_format = "0.0%"
        c_marg.alignment = Alignment(horizontal="center")

        c_val = ws3.cell(row=r_idx, column=11, value=total_contract)
        c_val.font = bold_font
        c_val.number_format = "$#,##0.00"

        ws3.cell(row=r_idx, column=12, value=vision_stat).font = regular_font

        c_stat = ws3.cell(row=r_idx, column=13, value=ccb_stat)
        c_stat.font = bold_font
        c_stat.alignment = Alignment(horizontal="center")
        if "APPROVED" in ccb_stat:
            c_stat.fill = green_alert_fill
        else:
            c_stat.fill = amber_alert_fill

        for col in range(1, 14):
            ws3.cell(row=r_idx, column=col).border = thin_border

    # Totals Row
    tot_row = 10
    ws3.cell(row=tot_row, column=1, value="Total Active Quotations").font = bold_font
    ws3.cell(row=tot_row, column=5, value="=SUM(E5:E9)").font = bold_font
    ws3.cell(row=tot_row, column=5).number_format = "$#,##0.00"
    ws3.cell(row=tot_row, column=6, value="=SUM(F5:F9)").font = bold_font
    ws3.cell(row=tot_row, column=6).number_format = "$#,##0.00"
    ws3.cell(row=tot_row, column=7, value="=SUM(G5:G9)").font = bold_font
    ws3.cell(row=tot_row, column=7).number_format = "$#,##0.00"
    ws3.cell(row=tot_row, column=8, value="=SUM(H5:H9)").font = bold_font
    ws3.cell(row=tot_row, column=8).number_format = "$#,##0.00"
    ws3.cell(row=tot_row, column=9, value="=SUM(I5:I9)").font = bold_font
    ws3.cell(row=tot_row, column=9).number_format = "$#,##0.00"
    ws3.cell(row=tot_row, column=10, value="=I10/K10").font = bold_font
    ws3.cell(row=tot_row, column=10).number_format = "0.0%"
    ws3.cell(row=tot_row, column=10).alignment = Alignment(horizontal="center")
    ws3.cell(row=tot_row, column=11, value="=SUM(K5:K9)").font = bold_font
    ws3.cell(row=tot_row, column=11).number_format = "$#,##0.00"

    for col in range(1, 14):
        ws3.cell(row=tot_row, column=col).border = bottom_double_border

    # ---------------------------------------------------------
    # TAB 4: PROJECT EXECUTION & FINANCIAL CONTROL
    # ---------------------------------------------------------
    ws4 = wb.create_sheet(title="Project Execution & Finance")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:L1")
    ws4["A1"] = "PROJECT EXECUTION & FINANCIAL MARGIN CONTROL REGISTRY"
    ws4["A1"].font = title_font
    ws4["A1"].fill = navy_fill
    ws4["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws4.row_dimensions[1].height = 26

    headers_proj = [
        "Project ID", "Client Name", "Project Title", "Contract Value ($)",
        "Invoiced to Date ($)", "Collected to Date ($)", "Outstanding Balance ($)",
        "Actual Cost to Date ($)", "Earned Progress (%)", "Current Profit ($)",
        "Current Margin (%)", "Margin Alert Status"
    ]
    for c_idx, h in enumerate(headers_proj, 1):
        cell = ws4.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[4].height = 22

    projects_data = [
        ("PRJ-2026-001", "ZimRe Real Estate Investment Trust", "Duplex Extension & Alterations", 45000.00, 18000.00, 18000.00, 14800.00, 0.40, "HEALTHY MARGIN (17.8%)"),
        ("PRJ-2026-002", "Graniteside Logistics Park", "Boundary Wall & Paving Package", 25000.00, 25000.00, 22500.00, 20400.00, 1.00, "COMPLETED (18.4%)"),
        ("PRJ-2026-003", "Old Mutual Property Zimbabwe", "Commercial Office Prep", 120000.00, 30000.00, 30000.00, 24500.00, 0.25, "HEALTHY MARGIN (18.3%)"),
    ]

    for r_idx, p in enumerate(projects_data, 5):
        pid, client, title, contract_val, invoiced, collected, actual_cost, progress, alert = p

        ws4.cell(row=r_idx, column=1, value=pid).font = bold_font
        ws4.cell(row=r_idx, column=2, value=client).font = regular_font
        ws4.cell(row=r_idx, column=3, value=title).font = regular_font

        c_val = ws4.cell(row=r_idx, column=4, value=contract_val)
        c_val.font = bold_font
        c_val.number_format = "$#,##0.00"

        c_inv = ws4.cell(row=r_idx, column=5, value=invoiced)
        c_inv.font = regular_font
        c_inv.number_format = "$#,##0.00"

        c_col = ws4.cell(row=r_idx, column=6, value=collected)
        c_col.font = regular_font
        c_col.number_format = "$#,##0.00"

        # Outstanding Balance Formula
        c_out = ws4.cell(row=r_idx, column=7, value=f"=D{r_idx}-F{r_idx}")
        c_out.font = bold_font
        c_out.number_format = "$#,##0.00"

        c_act = ws4.cell(row=r_idx, column=8, value=actual_cost)
        c_act.font = regular_font
        c_act.number_format = "$#,##0.00"

        c_prog = ws4.cell(row=r_idx, column=9, value=progress)
        c_prog.font = bold_font
        c_prog.number_format = "0.0%"
        c_prog.alignment = Alignment(horizontal="center")

        # Current Profit Formula
        c_prof = ws4.cell(row=r_idx, column=10, value=f"=E{r_idx}-H{r_idx}")
        c_prof.font = bold_font
        c_prof.number_format = "$#,##0.00"

        # Current Margin Formula
        c_marg = ws4.cell(row=r_idx, column=11, value=f"=J{r_idx}/E{r_idx}")
        c_marg.font = bold_font
        c_marg.number_format = "0.0%"
        c_marg.alignment = Alignment(horizontal="center")

        c_alert = ws4.cell(row=r_idx, column=12, value=alert)
        c_alert.font = bold_font
        c_alert.alignment = Alignment(horizontal="center")
        if "HEALTHY" in alert or "COMPLETED" in alert:
            c_alert.fill = green_alert_fill
        else:
            c_alert.fill = red_alert_fill

        for col in range(1, 13):
            ws4.cell(row=r_idx, column=col).border = thin_border

    # ---------------------------------------------------------
    # TAB 5: VERIFIED MATERIAL SOURCING & VENDOR MATRIX
    # ---------------------------------------------------------
    ws5 = wb.create_sheet(title="Verified Material Sourcing")
    ws5.views.sheetView[0].showGridLines = True

    ws5.merge_cells("A1:J1")
    ws5["A1"] = "VERIFIED MATERIAL SUPPLIER & PROCUREMENT MATRIX (WHERE TO BUY)"
    ws5["A1"].font = title_font
    ws5["A1"].fill = navy_fill
    ws5["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws5.row_dimensions[1].height = 26

    headers_sourcing = [
        "Package", "Material Item Description", "Unit", "Target Rate ($)",
        "Verified Supplier Name", "Physical Depot Address", "Contact Phone",
        "Lead Time (Days)", "Stock Availability", "CCB Quality Gate"
    ]
    for c_idx, h in enumerate(headers_sourcing, 1):
        cell = ws5.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[4].height = 22

    sourcing_data = [
        ("Substructure", "Cement 42.5N (50kg bags)", "bag", 12.50, "Apex Building Supplies", "14 Plymouth Rd, Southerton, Harare", "+263 242 754900", 2, "In Stock (500+ bags)", "Strict batching & bag count gate"),
        ("Substructure", "Concrete River Sand", "m3", 22.00, "Harare Aggregates", "Plot 4 Pomeroy Quarry, Mutare Rd, Msasa", "+263 772 100200", 1, "In Stock (Silt < 3%)", "Check silt content before unloading"),
        ("Substructure", "19mm Crushed Stone", "m3", 28.00, "Pomeroy Quarries", "Msasa Industrial Park, Harare", "+263 242 486711", 1, "In Stock (Weighbridge verified)", "Verify load volume via weighbridge"),
        ("Masonry", "Standard Common Bricks", "pcs", 0.18, "Willdale Bricks Harare", "Mt Hampden Brickworks, Lomagundi Rd", "+263 242 334600", 3, "Available (Order Week 2)", "Max 5% wastage allowance on site"),
        ("Masonry", "Brickforce Reinforcement 150mm", "m", 0.85, "Steelnet Zimbabwe", "100 Cripps Rd, Graniteside, Harare", "+263 242 751480", 2, "In Stock", "Laps of 300mm enforced"),
        ("Roofing", "0.47mm IBR Galvanised Sheeting", "m2", 13.00, "Chromadek Zimbabwe", "Coventry Rd, Workington, Harare", "+263 242 661200", 5, "Pre-order Week 3", "Store under waterproof tarpaulins"),
        ("Roofing", "SA Pine Structural Timber Trusses", "m2", 11.00, "Board & Timber Co", "Lytton Rd, Workington, Harare", "+263 242 757800", 4, "Custom Prep (4 days)", "Inspect timber grade and treatment"),
        ("Envelope", "Aluminium Window & Door Package", "sum", 4800.00, "AluSpec Zimbabwe", "12 Seke Rd, Graniteside, Harare", "+263 773 400500", 7, "Fabrication Lead Time 7d", "Measure opening dimensions on site"),
        ("Finishes", "Porcelain Floor Tiles 600x600", "m2", 14.00, "Tile Centre Msasa", "Enterprise Rd, Harare", "+263 242 487900", 5, "In Stock (Batch matched)", "Order 10% extra for cuts & spares"),
        ("MEP", "Electrical Cable & DB Component Pack", "sum", 3800.00, "Electric Centre", "88 Cameron St, Harare CBD", "+263 242 770100", 3, "SAZ Certified Stock", "Require SAZ compliance certs"),
    ]

    for r_idx, s_item in enumerate(sourcing_data, 5):
        pkg, mat, unit, rate, vendor, addr, phone, lead, stock, gate = s_item

        ws5.cell(row=r_idx, column=1, value=pkg).font = bold_font
        ws5.cell(row=r_idx, column=2, value=mat).font = regular_font
        ws5.cell(row=r_idx, column=3, value=unit).font = regular_font
        ws5.cell(row=r_idx, column=3).alignment = Alignment(horizontal="center")

        c_rate = ws5.cell(row=r_idx, column=4, value=rate)
        c_rate.font = bold_font
        c_rate.number_format = "$#,##0.00"

        ws5.cell(row=r_idx, column=5, value=vendor).font = bold_font
        ws5.cell(row=r_idx, column=6, value=addr).font = regular_font
        ws5.cell(row=r_idx, column=7, value=phone).font = regular_font

        c_lead = ws5.cell(row=r_idx, column=8, value=lead)
        c_lead.font = regular_font
        c_lead.alignment = Alignment(horizontal="center")

        ws5.cell(row=r_idx, column=9, value=stock).font = regular_font
        ws5.cell(row=r_idx, column=10, value=gate).font = regular_font

        for col in range(1, 11):
            ws5.cell(row=r_idx, column=col).border = thin_border

    # ---------------------------------------------------------
    # TAB 6: SUBCONTRACTOR PERFORMANCE SCORECARDS
    # ---------------------------------------------------------
    ws6 = wb.create_sheet(title="Subcontractor Scorecards")
    ws6.views.sheetView[0].showGridLines = True

    ws6.merge_cells("A1:I1")
    ws6["A1"] = "SUBCONTRACTOR PERFORMANCE & RATE INTELLIGENCE REGISTRY"
    ws6["A1"].font = title_font
    ws6["A1"].fill = navy_fill
    ws6["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws6.row_dimensions[1].height = 26

    headers_subby = [
        "Subby ID", "Trade Specialty", "Company Name", "Contact Person",
        "Phone Number", "CCB Score Rating (1-5)", "Rate Outlier Status",
        "Completed Projects", "HSE & Quality Compliance"
    ]
    for c_idx, h in enumerate(headers_subby, 1):
        cell = ws6.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[4].height = 22

    subby_data = [
        ("SUB-101", "Substructure & Excavation", "EarthMov Zimbabwe", "Jacob Moyo", "+263 772 331122", 4.8, "APPROVED RATE", 12, "Fully Compliant (ISO 45001)"),
        ("SUB-102", "Structural Bricklaying", "Master Masonry Civils", "Tafadzwa Chigumba", "+263 773 442233", 4.6, "APPROVED RATE", 18, "Fully Compliant"),
        ("SUB-103", "Timber Truss & Roofing", "Apex Timber Framing", "Craig Anderson", "+263 771 553344", 4.9, "APPROVED RATE", 14, "Fully Compliant"),
        ("SUB-104", "Plastering & Wet Trades", "Premier Renderers", "Kelvin Phiri", "+263 774 664455", 3.2, "RATE OUTLIER (+8% HIGH)", 6, "Conditional Approval"),
        ("SUB-105", "Electrical Reticulation", "BrightSpark Engineering", "Engineer Musarurwa", "+263 775 775566", 4.7, "APPROVED RATE", 22, "SAZ Certified Electricians"),
        ("SUB-106", "Plumbing & Sanitaryware", "FlowTech Plumbing", "Givemore Zvobgo", "+263 776 886677", 4.5, "APPROVED RATE", 15, "Fully Compliant"),
    ]

    for r_idx, sub in enumerate(subby_data, 5):
        sid, trade, comp, contact, phone, rating, outlier, projects, comp_stat = sub

        ws6.cell(row=r_idx, column=1, value=sid).font = bold_font
        ws6.cell(row=r_idx, column=2, value=trade).font = regular_font
        ws6.cell(row=r_idx, column=3, value=comp).font = bold_font
        ws6.cell(row=r_idx, column=4, value=contact).font = regular_font
        ws6.cell(row=r_idx, column=5, value=phone).font = regular_font

        c_rate = ws6.cell(row=r_idx, column=6, value=rating)
        c_rate.font = bold_font
        c_rate.alignment = Alignment(horizontal="center")

        c_out = ws6.cell(row=r_idx, column=7, value=outlier)
        c_out.font = bold_font
        c_out.alignment = Alignment(horizontal="center")
        if "APPROVED" in outlier:
            c_out.fill = green_alert_fill
        else:
            c_out.fill = red_alert_fill

        c_proj = ws6.cell(row=r_idx, column=8, value=projects)
        c_proj.font = regular_font
        c_proj.alignment = Alignment(horizontal="center")

        ws6.cell(row=r_idx, column=9, value=comp_stat).font = regular_font

        for col in range(1, 10):
            ws6.cell(row=r_idx, column=col).border = thin_border

    # ---------------------------------------------------------
    # TAB 7: CCB COMMERCIAL GUARD AUDIT LOG
    # ---------------------------------------------------------
    ws7 = wb.create_sheet(title="CCB Guard Audits")
    ws7.views.sheetView[0].showGridLines = True

    ws7.merge_cells("A1:J1")
    ws7["A1"] = "CCB COMMERCIAL GUARD & INTERCEPTION AUDIT EVIDENCE REGISTER"
    ws7["A1"].font = title_font
    ws7["A1"].fill = navy_fill
    ws7["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws7.row_dimensions[1].height = 26

    headers_audits = [
        "Audit ID", "Timestamp", "System Module", "Item / Request Description",
        "Requested Amount ($)", "CCB Baseline Ceiling ($)", "Discrepancy ($)",
        "Variance (%)", "System Action Taken", "Governance Status"
    ]
    for c_idx, h in enumerate(headers_audits, 1):
        cell = ws7.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[4].height = 22

    audits_data = [
        ("AUD-901", "2026-07-22 14:15", "Procurement PO Interceptor", "Highland Park Retail Glazing Package", 5200.00, 4800.00, "PO Blocked (Rate Outlier)", "REJECTED BY CCB"),
        ("AUD-902", "2026-07-22 15:30", "Site Operations Sync", "Site Concrete Request (Duplex Slab)", 440.00, 440.00, "Earned Qty Sync Verified", "PASSED AUDIT"),
        ("AUD-903", "2026-07-21 11:05", "Procurement PO Interceptor", "Plastering Labour Subcontractor Rate", 18.00, 15.00, "Subby Rate Outlier Blocked", "REJECTED BY CCB"),
        ("AUD-904", "2026-07-20 09:45", "Executive Governance Portal", "Budget Baseline Lock (SNC-RES-45K-001)", 45000.00, 45000.00, "Contract Baseline Locked", "PASSED AUDIT"),
        ("AUD-905", "2026-07-19 16:20", "Procurement PO Interceptor", "Cement 42.5N Emergency Procurement", 14.50, 12.50, "Supplier Rate Outlier Blocked", "REJECTED BY CCB"),
        ("AUD-906", "2026-07-18 10:12", "Site Operations Sync", "Site Brick Material Request (18.5k Bricks)", 3330.00, 3330.00, "Quantity Audit Passed", "PASSED AUDIT"),
        ("AUD-907", "2026-07-17 13:40", "Procurement PO Interceptor", "Unapproved Timber Truss Delivery Note", 3200.00, 2750.00, "PO Blocked (No Variation Order)", "REJECTED BY CCB"),
        ("AUD-908", "2026-07-16 08:50", "Site Operations Sync", "Rebar Y16 Site Drawdown Request", 850.00, 850.00, "Progress Baseline Matched", "PASSED AUDIT"),
    ]

    for r_idx, aud in enumerate(audits_data, 5):
        aid, ts, module, desc, req_amt, ceil_amt, action, gov = aud

        ws7.cell(row=r_idx, column=1, value=aid).font = bold_font
        ws7.cell(row=r_idx, column=2, value=ts).font = regular_font
        ws7.cell(row=r_idx, column=3, value=module).font = regular_font
        ws7.cell(row=r_idx, column=4, value=desc).font = regular_font

        c_req = ws7.cell(row=r_idx, column=5, value=req_amt)
        c_req.font = bold_font
        c_req.number_format = "$#,##0.00"

        c_ceil = ws7.cell(row=r_idx, column=6, value=ceil_amt)
        c_ceil.font = regular_font
        c_ceil.number_format = "$#,##0.00"

        # Discrepancy Formula
        c_disc = ws7.cell(row=r_idx, column=7, value=f"=E{r_idx}-F{r_idx}")
        c_disc.font = bold_font
        c_disc.number_format = "$#,##0.00"

        # Variance % Formula
        c_var = ws7.cell(row=r_idx, column=8, value=f"=G{r_idx}/F{r_idx}")
        c_var.font = bold_font
        c_var.number_format = "0.0%"
        c_var.alignment = Alignment(horizontal="center")

        ws7.cell(row=r_idx, column=9, value=action).font = regular_font

        c_gov = ws7.cell(row=r_idx, column=10, value=gov)
        c_gov.font = bold_font
        c_gov.alignment = Alignment(horizontal="center")
        if "PASSED" in gov:
            c_gov.fill = green_alert_fill
        else:
            c_gov.fill = red_alert_fill

        for col in range(1, 11):
            ws7.cell(row=r_idx, column=col).border = thin_border

    # ---------------------------------------------------------
    # AUTO-FIT COLUMN WIDTHS ACROSS ALL WORKSHEETS
    # ---------------------------------------------------------
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Avoid merged title row skewing column width
                if cell.row in (1, 2) and col_letter in ("A", "B", "C"):
                    continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    DESKTOP_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)

    wb.save(str(DESKTOP_EXCEL_PATH))
    shutil.copy(str(DESKTOP_EXCEL_PATH), str(ARTIFACT_EXCEL_PATH))
    return DESKTOP_EXCEL_PATH


if __name__ == "__main__":
    path = create_crm_workbook()
    print(f"EXCEL_CREATION_SUCCESS: {path}")
